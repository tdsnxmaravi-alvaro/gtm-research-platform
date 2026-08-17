"""Tests for Phase 3: consolidate (master) + outreach (.eml)."""

import email
import email.policy

from gtm.config import CampaignConfig
from gtm.consolidate import normalize_name, build_master
from gtm.outreach import render_template, write_eml
from gtm.outreach.eml import apply_template


def _cfg(**over):
    d = dict(name="p3", target_type="resellers", mode="provided", country="Spain",
             products=[{"name": "Trimble", "value_prop": "design software",
                        "fit_criteria": ["sells software"]}],
             provided_list_path="x.csv",
             outreach={"enabled": True, "language": "es", "min_tier": "B",
                       "sender_name": "Ana BDR", "sender_email": "ana@tdsynnex.com"})
    d.update(over)
    return CampaignConfig(**d)


def test_normalize_name():
    assert normalize_name("SERVICIOS, S.A.") == "SERVICIOS"
    assert normalize_name("Acme Inc.") == "ACME"
    assert normalize_name("Foo (Bar) LLC") == "FOO"


def test_build_master_neutralizes_formula_injection(tmp_path):
    import csv
    out = tmp_path / "camp"
    out.mkdir()
    formula = '=HYPERLINK("http://evil.example","x")'
    with open(out / "results.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "company", "website", "final_tier", "score", "fit_summary",
            "recommended_products", "evidence_urls", "product"])
        w.writeheader()
        w.writerow({
            "company": "Acme SA", "website": "https://a.es", "final_tier": "A",
            "score": "88", "fit_summary": formula, "recommended_products": "SketchUp",
            "evidence_urls": "https://a.es/x", "product": "Trimble",
        })
    (out / "contacts.csv").write_text(
        "company,contact_name,title,email,email_status,direct_phone,linkedin\n"
        "Acme SA,Ana Ruiz,CEO,ana@a.es,verified,,\n",
        encoding="utf-8-sig")
    build_master(_cfg(), out_dir=out, min_tier="D")
    csv_text = (out / "master.csv").read_text(encoding="utf-8-sig")
    assert "'=HYPERLINK" in csv_text
    from openpyxl import load_workbook
    wb = load_workbook(out / "master.xlsx")
    ws = wb["Master Outreach"]
    headers = [c.value for c in ws[1]]
    fit_col = headers.index("Fit Summary") + 1
    cell = ws.cell(row=2, column=fit_col)
    assert str(cell.value).startswith("'=")
    assert cell.data_type != "f"


def test_build_master_joins_results_and_contacts(tmp_path):
    out = tmp_path / "camp"
    out.mkdir()
    (out / "results.csv").write_text(
        "company,website,final_tier,score,fit_summary,recommended_products,evidence_urls,product\n"
        "Acme SA,https://a.es,A,88,strong fit,SketchUp,https://a.es/x,Trimble\n"
        "Beta SL,https://b.es,C,55,weak,SketchUp,https://b.es/y,Trimble\n",
        encoding="utf-8-sig")
    (out / "contacts.csv").write_text(
        "company,contact_name,title,email,email_status,direct_phone,linkedin\n"
        "Acme SA,Ana Ruiz,CEO,ana@a.es,verified,,\n"
        "Acme SA,Bob Diaz,CTO,bob@a.es,verified,,\n",
        encoding="utf-8-sig")

    c = _cfg()
    rows = build_master(c, out_dir=out, min_tier="B")
    # Beta (C) excluded by min_tier B; Acme has 2 contacts
    assert len(rows) == 2
    assert all(r["company"] == "Acme SA" for r in rows)
    assert rows[0]["tier"] == "A"
    assert (out / "master.csv").exists()
    assert (out / "master.xlsx").exists()


def test_build_master_merges_domain_variants_and_dedupes_contacts(tmp_path):
    out = tmp_path / "camp"
    out.mkdir()
    # Two name variants of the same site (trimech.com) + a duplicate contact row.
    (out / "results.csv").write_text(
        "company,website,final_tier,score,fit_summary,recommended_products,evidence_urls,product\n"
        "TriMech,https://trimech.com/,A,99,fit,X,https://trimech.com/x,Unity\n"
        "TriMech Group / Javelin,https://trimech.com/,A,94,fit,X,https://trimech.com/y,Unity\n",
        encoding="utf-8-sig")
    (out / "contacts.csv").write_text(
        "company,website,contact_name,title,email,email_status,direct_phone,linkedin\n"
        "TriMech,trimech.com,Craig Oznick,VP,craig@trimech.com,verified,,\n"
        "TriMech,trimech.com,Craig Oznick,VP,craig@trimech.com,verified,,\n",
        encoding="utf-8-sig")

    rows = build_master(_cfg(), out_dir=out, min_tier="D")
    # Both variants merge into ONE company (highest score kept), and the duplicate
    # contact collapses to a single row.
    assert len(rows) == 1
    assert rows[0]["company"] == "TriMech"
    assert rows[0]["score"] == "99"
    assert rows[0]["email"] == "craig@trimech.com"


def test_outreach_auto_localizes_per_row_country():
    # No explicit outreach language + empty campaign country => auto per-row country.
    c = CampaignConfig(name="loc", target_type="resellers", mode="provided",
                       country="", vendor="Trimble",
                       products=[{"name": "Trimble", "value_prop": "AEC"}],
                       provided_list_path="x.csv", outreach={"enabled": True})
    assert c.outreach.language is None
    _, es = render_template(c, {"company": "A", "contact_name": "Pablo", "country": "Spain"})
    _, pt = render_template(c, {"company": "A", "contact_name": "Pablo", "country": "Portugal"})
    _, en = render_template(c, {"company": "A", "contact_name": "Pablo", "country": ""})
    assert es.startswith("Hola") and pt.startswith("Olá") and en.startswith("Hi")


def test_render_template_es():
    c = _cfg()
    subj, body = render_template(c, {"company": "Acme SA", "contact_name": "Ana Ruiz",
                                     "recommended_products": "SketchUp", "product": "Trimble",
                                     "fit_summary": "encaja bien"})
    assert "Acme SA" in subj
    assert body.startswith("Hola Ana")
    assert "Trimble" in body
    assert body.rstrip().endswith("Un saludo,")  # consistent closing, no name/org
    assert "Ana BDR" not in body  # identity comes from the Outlook signature


def test_signoff_normalized_across_variants():
    from gtm.outreach.email_gen import _apply_signoff

    c = CampaignConfig(name="x", target_type="resellers", mode="provided", country="",
                       vendor="Trimble", products=[{"name": "Trimble", "value_prop": "AEC"}],
                       provided_list_path="x.csv", outreach={"enabled": True})
    variants = [
        "Hi Rohit,\n\nGreat fit.\n\nBest,\nTD SYNNEX team\nTD SYNNEX",
        "Hi Rohit,\n\nGreat fit.\n\nTD SYNNEX",
        "Hi Rohit,\n\nGreat fit.\n\nBest,\nTD SYNNEX",
    ]
    for v in variants:
        assert _apply_signoff(v, "en", c).endswith("Great fit.\n\nBest regards,")


def test_branded_frame_has_editable_line_above_card():
    from gtm.outreach.eml import _branded_html, _plain_to_html

    for html in (_branded_html("Hi,\n\nBody."), _plain_to_html("Hi,\n\nBody.")):
        # Editable spacer paragraphs sit above AND below the branded card so
        # Outlook's signature lands above it and Enter at the bottom extends
        # outside the box, not inside the card.
        card = html.index("max-width:640px")
        assert html.index("&nbsp;</p>") < card < html.rindex("&nbsp;</p>")


def test_write_eml_strips_crlf_from_subject(tmp_path):
    path = write_eml(
        tmp_path / "inj.eml",
        to_email="ana@a.es",
        to_name="Ana",
        subject="Hola\r\nBcc: evil@x.com",
        body="Hola Ana\n\nSaludos",
        from_email="bdr@tdsynnex.com",
        from_name="BDR",
    )
    with open(path, "rb") as f:
        msg = email.message_from_binary_file(f, policy=email.policy.default)
    assert msg.get("Bcc") is None
    assert "\r" not in (msg["Subject"] or "") and "\n" not in (msg["Subject"] or "")
    assert msg["Subject"] == "Hola Bcc: evil@x.com"


def test_write_eml_is_smtp_and_unsent(tmp_path):
    path = write_eml(tmp_path / "d.eml", to_email="ana@a.es", to_name="Ana",
                     subject="Hola", body="Hola Ana\n\nSaludos",
                     from_email="bdr@tdsynnex.com", from_name="BDR")
    raw = path.read_bytes()
    # SMTP policy -> CRLF line endings (avoids Outlook quoted-printable corruption)
    assert b"\r\n" in raw
    with open(path, "rb") as f:
        msg = email.message_from_binary_file(f, policy=email.policy.default)
    assert msg["X-Unsent"] == "1"
    assert msg["Subject"] == "Hola"
    assert msg.is_multipart()  # text + html alternative


def test_apply_template_replaces_marker():
    html = '<div class="box"><img src="cid:banner"><div>{{BODY}}</div>-- TD SYNNEX</div>'
    out = apply_template(html, "Hola Acme.\n\nSaludos")
    assert "{{BODY}}" not in out
    assert "Hola Acme." in out and "Saludos" in out
    assert 'class="box"' in out and "TD SYNNEX" in out  # box preserved
    assert apply_template("<div>no marker</div>", "x") is None


def test_write_eml_uses_branded_template_and_carries_inline_image(tmp_path):
    # Build a branded sample .eml with a {{BODY}} marker + inline CID image.
    from email.message import EmailMessage
    tpl = EmailMessage()
    tpl["Subject"], tpl["From"], tpl["To"] = "s", "a@b.com", "c@d.com"
    tpl.set_content("plain")
    tpl.add_alternative(
        '<html><body><div style="border:1px solid #ccc">'
        '<img src="cid:banner1"><div>{{BODY}}</div><div>-- TD SYNNEX</div>'
        "</div></body></html>",
        subtype="html",
    )
    tpl.get_payload()[-1].add_related(b"\x89PNG_fake", maintype="image",
                                      subtype="png", cid="<banner1>")
    sample = tmp_path / "sample.eml"
    sample.write_bytes(tpl.as_bytes(policy=email.policy.SMTP))

    out = write_eml(tmp_path / "out.eml", to_email="x@y.com", subject="Hi",
                    body="Hola Acme.\n\nEsto es una prueba.", template_eml=str(sample))
    msg = email.message_from_bytes(out.read_bytes(), policy=email.policy.default)
    html = next(p.get_content() for p in msg.walk() if p.get_content_type() == "text/html")
    imgs = [p for p in msg.walk() if p.get_content_maintype() == "image"]
    assert "{{BODY}}" not in html and "Hola Acme." in html
    assert "border:1px solid #ccc" in html and "cid:banner1" in html  # box + ref kept
    assert len(imgs) == 1 and imgs[0]["Content-ID"] == "<banner1>"  # inline image carried


def test_write_eml_with_logo_uses_frame_and_inline_banner(tmp_path):
    logo = tmp_path / "logo.png"
    logo.write_bytes(b"\x89PNG\r\n\x1a\n_fakelogo")
    out = write_eml(tmp_path / "o.eml", to_email="x@y.com", subject="Hi",
                    body="Hola Acme.\n\nSaludos", logo_path=str(logo))
    msg = email.message_from_bytes(out.read_bytes(), policy=email.policy.default)
    html = next(p.get_content() for p in msg.walk() if p.get_content_type() == "text/html")
    imgs = [p for p in msg.walk() if p.get_content_maintype() == "image"]
    assert "<img" in html and "cid:" in html and "Hola Acme." in html
    assert "border:1px solid #e1e4e8" in html  # branded frame
    assert len(imgs) == 1
    assert imgs[0]["Content-ID"].strip("<>") in html  # banner cid matches


def test_run_outreach_writes_eml(tmp_path, monkeypatch):
    from pathlib import Path
    from gtm.outreach.runner import run_outreach

    out = tmp_path / "camp"
    out.mkdir()
    (out / "master.csv").write_text(
        "company,tier,score,contact_name,title,email,direct_phone\n"
        "Acme SA,A,90,Ana Ruiz,CEO,ana@a.es,\n",
        encoding="utf-8-sig")

    monkeypatch.setattr(
        "gtm.outreach.runner.generate_outreach",
        lambda *_a, **_k: {
            "subject": "Hi", "body": "Hello",
            "followup_subject": "", "followup_body": "", "talking_points": "",
        },
    )

    def _write(path, **_kw):
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("From: x\nTo: y\nSubject: Hi\n\nHello\n", encoding="utf-8")
        return path

    monkeypatch.setattr("gtm.outreach.runner.write_eml", _write)
    monkeypatch.setattr("gtm.consolidate.master.write_outreach_sheet", lambda *_a, **_k: None)

    drafts = run_outreach(_cfg(), out_dir=out, use_agent=False)
    assert len(drafts) == 1
    assert drafts[0]["email"] == "ana@a.es"
    assert Path(drafts[0]["eml"]).exists()
