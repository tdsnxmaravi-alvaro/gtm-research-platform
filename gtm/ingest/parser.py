"""Ingest — parse LLM research output and provided input lists into normalized rows.

- parse_results(text): robustly extract the fixed JSON schema {"results": [...]}
  from an LLM response (handles markdown fences / surrounding prose) and normalize
  each result to standard columns, including evidence-URL bookkeeping.
- load_provided_list(path): read a supplied company list (CSV) with normalized headers.
- write_rows_csv(rows, path): persist normalized rows.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

RESULT_COLUMNS = [
    "company", "website", "fit_summary", "score", "tier",
    "recommended_products", "notes",
    "evidence_urls", "evidence_count", "has_verified_url", "evidence",
]

# Header variants -> normalized name for provided input lists
_HEADER_MAP = {
    "company": "company", "company name": "company", "name": "company",
    "account": "company", "reseller": "company",
    "reseller name": "company", "resellername": "company", "partner name": "company",
    "website": "website", "url": "website", "domain": "website", "web": "website",
}


def _extract_json(text: str) -> dict | list | None:
    """Pull the outermost JSON object/array from a possibly-noisy LLM response.

    Prefers a JSON *object* over an array so that leading bracket noise (e.g.
    LARA `[[LARA_TOOL_ACTIVITY:...]]` markers) does not hijack the match.
    """
    text = text or ""
    candidates: list[str] = []
    fence = re.search(r"```(?:json)?\s*(\{.*\}|\[.*\])\s*```", text, re.DOTALL)
    if fence:
        candidates.append(fence.group(1))
    obj = re.search(r"\{.*\}", text, re.DOTALL)
    if obj:
        candidates.append(obj.group(0))
    arr = re.search(r"\[.*\]", text, re.DOTALL)
    if arr:
        candidates.append(arr.group(0))
    for cand in candidates:
        try:
            return json.loads(cand)
        except json.JSONDecodeError:
            continue
    return None


def normalize_result(r: dict) -> dict:
    """Normalize one LLM result object to the standard result columns.

    Supports the rubric format (`dimension_scores`: per-dimension points + evidence,
    summed deterministically) and the legacy holistic format (`score` + `evidence`).
    """
    dim_scores = r.get("dimension_scores")
    urls: list[str] = []
    breakdown: list[dict] = []
    score = r.get("score", "")

    if isinstance(dim_scores, list) and dim_scores:
        total = 0
        for d in dim_scores:
            if not isinstance(d, dict):
                continue
            try:
                pts = int(d.get("points", 0) or 0)
            except (TypeError, ValueError):
                pts = 0
            try:
                mx = int(d.get("max", 0) or 0)
            except (TypeError, ValueError):
                mx = 0
            if mx > 0:
                pts = max(0, min(pts, mx))  # clamp to the dimension's max
            total += pts
            url = str(d.get("evidence_url") or "").strip()
            if url:
                urls.append(url)
            breakdown.append({"name": str(d.get("name", "")).strip(),
                              "points": pts, "max": mx,
                              "rationale": str(d.get("rationale", "")).strip(),
                              "evidence_url": url})
        score = total
        evidence_blob = breakdown
    else:
        evidence = r.get("evidence") or []
        if not isinstance(evidence, list):
            evidence = []
        urls = [str(e.get("url")).strip() for e in evidence
                if isinstance(e, dict) and e.get("url")]
        evidence_blob = evidence

    # De-dup URLs preserving order
    seen: list[str] = []
    for u in urls:
        if u and u not in seen:
            seen.append(u)
    urls = seen

    rec = r.get("recommended_products") or []
    if isinstance(rec, str):
        rec = [rec]
    return {
        "company": str(r.get("company", "")).strip(),
        "website": str(r.get("website", "")).strip(),
        "fit_summary": str(r.get("fit_summary", "")).strip(),
        "score": score,
        "tier": str(r.get("tier", "")).strip().upper(),
        "recommended_products": "; ".join(str(x) for x in rec),
        "notes": str(r.get("notes", "")).strip(),
        "evidence_urls": "; ".join(urls),
        "evidence_count": len(urls),
        "has_verified_url": bool(urls),
        "evidence": json.dumps(evidence_blob, ensure_ascii=False),
    }


def parse_results(text: str) -> list[dict]:
    """Parse an LLM research response into a list of normalized result rows."""
    obj = _extract_json(text)
    if obj is None:
        return []
    results = obj.get("results") if isinstance(obj, dict) else obj
    if not isinstance(results, list):
        return []
    return [normalize_result(r) for r in results if isinstance(r, dict)]


def load_provided_list(path: str | Path,
                       column_overrides: dict[str, str] | None = None) -> list[dict]:
    """Read a provided company list and normalize its headers.

    Supports CSV and Excel (.xlsx). The first row is treated as the header; the
    first sheet is used for Excel. Header variants are mapped to `company` /
    `website` via `_HEADER_MAP`. `column_overrides` (raw-header-lowercased ->
    canonical name) takes precedence and is how the AI schema-mapper injects a
    smarter mapping for non-standard files.
    """
    path = Path(path)
    overrides = {(k or "").strip().lower(): v for k, v in (column_overrides or {}).items()}
    raw_rows = _read_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm") else _read_csv(path)
    rows: list[dict] = []
    for raw in raw_rows:
        row: dict = {}
        for k, v in raw.items():
            low = (k or "").strip().lower()
            key = overrides.get(low) or _HEADER_MAP.get(low, low)
            row[key] = ("" if v is None else str(v)).strip()
        if row.get("company"):
            rows.append(row)
    return rows


def _read_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []
    out: list[dict] = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        out.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
    wb.close()
    return out


# Non-key columns commonly useful as qualification context (optional, may be absent).
_CONTEXT_HINT_FIELDS = {
    "other software in use", "company size", "number of employees",
    "end customer country", "country", "industry", "sector",
}


def _raw_headers(path: Path) -> list[str]:
    """Return the header row of a CSV/XLSX provided list, in order."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        try:
            row = next(ws.iter_rows(values_only=True))
        except StopIteration:
            row = ()
        wb.close()
        return [str(c).strip() if c is not None else "" for c in row]
    with open(path, encoding="utf-8-sig", newline="") as f:
        return next(csv.reader(f), [])


def inspect_provided_list(path: str | Path, use_ai: bool = False) -> dict:
    """Pre-flight a provided list: detect columns, header mapping, and data quality.

    Returns a report dict (headers, mapping, counts, context fields, warnings, ok).
    Use this before a run to confirm the minimum required fields (company + website).

    When `use_ai` is True and a schema-mapper agent is configured, an AI mapping
    is added (and applied for column detection) — useful for non-standard files
    whose headers aren't in the built-in map. Only headers + non-PII samples are
    sent to the AI.
    """
    path = Path(path)
    fmt = "xlsx" if path.suffix.lower() in (".xlsx", ".xlsm") else "csv"
    headers = _raw_headers(path)
    raw = _read_xlsx(path) if fmt == "xlsx" else _read_csv(path)
    raw_total = len(raw)

    # Optional AI-assisted mapping (PII-minimized); falls back silently.
    ai_mapping = None
    overrides: dict[str, str] = {}
    if use_ai:
        from .schema_ai import ai_map_columns, overrides_from_ai
        ai_mapping = ai_map_columns(headers, raw)
        if ai_mapping:
            overrides = overrides_from_ai(ai_mapping)

    def _canonical(h: str) -> str:
        low = h.strip().lower()
        return overrides.get(low) or _HEADER_MAP.get(low, low)

    mapped = {h: _canonical(h) for h in headers if h}

    rows = load_provided_list(path, column_overrides=overrides)  # only rows with a company
    with_company = len(rows)
    with_website = sum(1 for r in rows if r.get("website"))

    seen: dict[str, int] = {}
    for r in rows:
        key = r["company"].strip().lower()
        seen[key] = seen.get(key, 0) + 1
    duplicates = sum(1 for v in seen.values() if v > 1)

    has_company_col = "company" in mapped.values()
    has_website_col = "website" in mapped.values()
    context_present = [h.strip().lower() for h in headers
                       if h.strip().lower() in _CONTEXT_HINT_FIELDS]

    warnings: list[str] = []
    if not has_company_col:
        warnings.append("No company/name column detected — this is REQUIRED.")
    if not has_website_col:
        warnings.append("No website/URL column detected — minimum recommended is name + website.")
    if raw_total and with_company < raw_total:
        warnings.append(f"{raw_total - with_company} row(s) have no company name and will be skipped.")
    if with_company and with_website / with_company < 0.5:
        warnings.append(
            f"Only {with_website}/{with_company} rows have a website — "
            "research/enrichment quality will drop without domains."
        )
    if duplicates:
        warnings.append(f"{duplicates} duplicate company name(s) detected.")
    if isinstance(ai_mapping, dict):
        for w in ai_mapping.get("warnings", []) or []:
            warnings.append(f"AI: {w}")

    return {
        "path": str(path),
        "format": fmt,
        "raw_headers": headers,
        "mapping": mapped,
        "ai_mapping": ai_mapping,
        "raw_rows": raw_total,
        "with_company": with_company,
        "with_website": with_website,
        "missing_website": with_company - with_website,
        "duplicates": duplicates,
        "context_fields_present": context_present,
        "has_company_col": has_company_col,
        "has_website_col": has_website_col,
        "warnings": warnings,
        "ok": has_company_col and with_company > 0,
    }


def write_rows_csv(rows: list[dict], path: str | Path, columns: list[str] | None = None) -> None:
    """Write normalized rows to CSV (utf-8-sig for Excel)."""
    if not rows:
        columns = columns or RESULT_COLUMNS
    else:
        columns = columns or list({k for r in rows for k in r}) or RESULT_COLUMNS
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
