"""Consolidate research results + enriched contacts into a master list.

One row per contact (or one per company when no contacts), carrying the company's
tier/score/evidence alongside the contact's email/phone. Sorted by tier then score,
deduplicated by normalized company name. Exports CSV and XLSX for the BDM.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

from ..config.schema import CampaignConfig

MASTER_COLS = [
    "tier", "score", "company", "website", "country", "employees", "software_resold",
    "datech_match", "datech_geo",
    "contact_name", "title", "email", "email_status",
    "direct_phone", "corporate_phone", "linkedin",
    "vendor", "recommended_products", "fit_summary", "evidence_urls",
]

# "Master Outreach" sheet — one row per company (header, source key).
SUMMARY_SHEET = [
    ("Company", "company"), ("Website", "website"), ("Tier", "tier"), ("Score", "score"),
    ("Country", "country"), ("Employees", "employees"), ("Software Resold", "software_resold"),
    ("Datech Match", "datech_match"), ("Datech Market", "datech_geo"),
    ("Independence", "independence"), ("Apollo Verified", "apollo_verified"),
    ("Validation", "validation"), ("Total Contacts", "total_contacts"),
    ("Verified Emails", "verified_emails"), ("Best Contact", "best_contact"),
    ("Best Contact Title", "best_title"), ("Best Contact Email", "best_email"),
    ("Best Contact Phone", "best_phone"), ("All Verified Emails", "all_verified_emails"),
    ("Vendor", "vendor"), ("Recommended Products", "recommended_products"),
    ("Fit Summary", "fit_summary"), ("Evidence URLs", "evidence_urls"),
]

# "All Contacts" sheet — one row per contact.
CONTACTS_SHEET = [
    ("Company", "company"), ("Tier", "tier"), ("Score", "score"),
    ("Contact Name", "contact_name"), ("Title", "title"), ("Email", "email"),
    ("Email Status", "email_status"), ("Direct Phone", "direct_phone"),
    ("Corporate Phone", "corporate_phone"), ("LinkedIn", "linkedin"),
    ("City", "city"), ("State", "state"), ("Country", "country"),
]

# "Outreach" sheet — one row per company (the ready-to-send copy). Talking points
# are only filled when a phone number was obtained for the contact.
OUTREACH_SHEET = [
    ("Company", "company"), ("Tier", "tier"), ("Score", "score"),
    ("Contact", "contact_name"), ("Title", "title"), ("Email", "email"),
    ("Phone", "phone"), ("Subject", "subject"), ("Body", "body"),
    ("Follow-up Subject", "followup_subject"), ("Follow-up Body", "followup_body"),
    ("Talking Points", "talking_points"),
]

_CONTACT_PRIORITY = ["president", "ceo", "owner", "founder", "managing director",
                     "director", "vp", "vice president", "head", "manager"]

_TIER_ORDER = {"A": 0, "B": 1, "C": 2, "D": 3, "": 9}


def normalize_name(name: str) -> str:
    """Normalize a company name for deduplication."""
    n = (name or "").upper().strip()
    n = re.sub(r"\(.*?\)", "", n)
    for suf in (" INC.", " INC", " LLC", " LTD.", " LTD", " CORP.", " CORP",
                " CO.", " CO", " L.P.", " LP", " LIMITED", " PTE",
                " S.L.U.", " S.L.", " SL", " S.A.", " SA", " SLU", " LDA"):
        if n.endswith(suf):
            n = n[:-len(suf)]
    n = re.sub(r"[^A-Z0-9 ]", "", n)
    return re.sub(r"\s+", " ", n).strip()


def _read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def build_master(config: CampaignConfig, out_dir: str | Path | None = None,
                 min_tier: str | None = None, progress_cb=None) -> list[dict]:
    """Join results.csv + contacts.csv into a master list. Returns master rows."""
    out = Path(out_dir or (Path("campaigns") / config.name))
    results = _read_csv(out / "results.csv")
    contacts = _read_csv(out / "contacts.csv")

    tier_cap = _TIER_ORDER.get((min_tier or "").upper(), 9)

    # Best result row per normalized company (keep highest score).
    best: dict[str, dict] = {}
    for r in results:
        key = normalize_name(r.get("company", ""))
        if not key:
            continue
        prev = best.get(key)
        if prev is None or _score(r) > _score(prev):
            best[key] = r

    _annotate_datech(config, best)

    # Group contacts by company.
    contacts_by_company: dict[str, list[dict]] = {}
    for c in contacts:
        key = normalize_name(c.get("company", ""))
        contacts_by_company.setdefault(key, []).append(c)

    rows: list[dict] = []
    summary: list[dict] = []
    contact_rows: list[dict] = []
    for key, r in best.items():
        tier = (r.get("final_tier") or r.get("tier") or "").upper()
        if min_tier and _TIER_ORDER.get(tier, 9) > tier_cap:
            continue
        cs = contacts_by_company.get(key, [])
        if cs:
            for c in cs:
                rows.append(_master_row(config, r, tier, c))
        else:
            rows.append(_master_row(config, r, tier, {}))
        contact_rows.extend(cs)
        summary.append(_summary_row(config, r, tier, cs))

    rows.sort(key=lambda x: (_TIER_ORDER.get(x["tier"], 9), -_num(x["score"])))
    summary.sort(key=lambda x: (_TIER_ORDER.get(x["tier"], 9), -_num(x["score"])))
    contact_rows.sort(key=lambda c: (_TIER_ORDER.get((c.get("tier") or "").upper(), 9),
                                     -_num(c.get("score"))))
    _write_csv(rows, out / "master.csv")
    _write_xlsx(summary, contact_rows, out / "master.xlsx")
    if progress_cb:
        progress_cb(len(rows), len(rows) or 1)
    print(f"Master: {len(summary)} companies, {len(rows)} contact rows -> "
          f"{out / 'master.csv'} (+ .xlsx: Master Outreach + All Contacts)")
    return rows


def _score(r: dict) -> float:
    return _num(r.get("score"))


def _annotate_datech(config: CampaignConfig, best: dict[str, dict]) -> None:
    """Flag existing Datech resellers (country-aware) if a list is configured."""
    path = getattr(config, "datech_reseller_list", None)
    if not path or not Path(path).exists():
        return
    from .datech_match import DatechIndex, load_datech_records
    index = DatechIndex([], records=load_datech_records(path))
    for r in best.values():
        m = index.match(r.get("company", ""), r.get("country") or config.country)
        r["datech_match"] = m["name"] if m else ""
        if m:
            where = m.get("country") or m.get("geo") or ""
            tag = {True: " (same market)", False: " (other market)"}.get(m.get("same_country"), "")
            r["datech_geo"] = (where + tag).strip()
        else:
            r["datech_geo"] = ""


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _master_row(config: CampaignConfig, r: dict, tier: str, c: dict) -> dict:
    return {
        "tier": tier,
        "score": r.get("score", ""),
        "company": r.get("company", ""),
        "website": r.get("website", ""),
        "country": r.get("country") or config.country,
        "employees": r.get("employees", ""),
        "software_resold": r.get("software_resold", ""),
        "datech_match": r.get("datech_match", ""),
        "datech_geo": r.get("datech_geo", ""),
        "contact_name": c.get("contact_name", ""),
        "title": c.get("title", ""),
        "email": c.get("email", ""),
        "email_status": c.get("email_status", ""),
        "direct_phone": c.get("direct_phone", ""),
        "corporate_phone": c.get("corporate_phone", ""),
        "linkedin": c.get("linkedin", ""),
        "vendor": config.vendor or r.get("product", ""),
        "recommended_products": r.get("recommended_products", ""),
        "fit_summary": r.get("fit_summary", ""),
        "evidence_urls": r.get("evidence_urls", ""),
    }


def _best_contact(cs: list[dict]) -> dict:
    """Pick the top-ranked contact that has an email (else the top-ranked)."""
    pool = [c for c in cs if c.get("email")] or cs
    if not pool:
        return {}

    def rank(c: dict) -> int:
        t = (c.get("title") or "").lower()
        return next((i for i, p in enumerate(_CONTACT_PRIORITY) if p in t), 99)

    return sorted(pool, key=rank)[0]


def _summary_row(config: CampaignConfig, r: dict, tier: str, cs: list[dict]) -> dict:
    """One company-level row for the 'Master Outreach' sheet."""
    verified = [c for c in cs
                if (c.get("email_status") or "").lower() == "verified" and c.get("email")]
    bc = _best_contact(cs)
    has_url = (str(r.get("has_verified_url", "")).lower() in ("true", "1", "yes")
               or _num(r.get("evidence_count")) > 0)
    return {
        "company": r.get("company", ""),
        "website": r.get("website", ""),
        "tier": tier,
        "score": r.get("score", ""),
        "country": r.get("country") or config.country,
        "employees": r.get("employees", ""),
        "software_resold": r.get("software_resold", ""),
        "datech_match": r.get("datech_match", ""),
        "datech_geo": r.get("datech_geo", ""),
        "independence": r.get("independence", ""),
        "apollo_verified": ("Yes" if any((c.get("email_status") or "").lower() == "verified"
                                          for c in cs) else ("No" if cs else "")),
        "validation": "PASS" if has_url else "REVIEW",
        "total_contacts": len(cs),
        "verified_emails": len(verified),
        "best_contact": bc.get("contact_name", ""),
        "best_title": bc.get("title", ""),
        "best_email": bc.get("email", ""),
        "best_phone": bc.get("direct_phone", "") or bc.get("corporate_phone", ""),
        "all_verified_emails": "; ".join(c.get("email", "") for c in verified),
        "vendor": config.vendor or r.get("product", ""),
        "recommended_products": r.get("recommended_products", ""),
        "fit_summary": r.get("fit_summary", ""),
        "evidence_urls": r.get("evidence_urls", ""),
    }


def _write_csv(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=MASTER_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(summary_rows: list[dict], contact_rows: list[dict], path: Path) -> None:
    """Write a two-sheet workbook: 'Master Outreach' (per company) + 'All Contacts'."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return
    wb = Workbook()
    _fill_sheet(wb.active, "Master Outreach", SUMMARY_SHEET, summary_rows)
    _fill_sheet(wb.create_sheet("All Contacts"), "All Contacts", CONTACTS_SHEET, contact_rows)
    try:
        wb.save(path)
    except PermissionError:
        print(f"  !! could not write {path.name} (is it open in Excel?). CSV written OK.")


def write_outreach_sheet(out_dir: str | Path, rows: list[dict]) -> None:
    """Add/refresh an 'Outreach' sheet on the campaign's master.xlsx.

    Keeps everything in one workbook (no extra Excel file). Called by the outreach
    stage after drafts are generated; talking points are only present for contacts
    whose phone number was obtained.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return
    path = Path(out_dir) / "master.xlsx"
    try:
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # drop the blank default sheet
    except (OSError, ValueError):
        wb = Workbook()
        wb.remove(wb.active)
    if "Outreach" in wb.sheetnames:
        wb.remove(wb["Outreach"])
    _fill_sheet(wb.create_sheet("Outreach"), "Outreach", OUTREACH_SHEET, rows)
    try:
        wb.save(path)
    except PermissionError:
        print(f"  !! could not write {path.name} (is it open in Excel?).")


_XLSX_WIDTHS = {
    "company": 30, "website": 26, "country": 12, "tier": 6, "score": 7,
    "employees": 12, "software_resold": 44, "independence": 14, "apollo_verified": 14,
    "validation": 11, "total_contacts": 13, "verified_emails": 14, "best_contact": 22,
    "best_title": 24, "best_email": 28, "best_phone": 16, "all_verified_emails": 40,
    "vendor": 14, "recommended_products": 32, "fit_summary": 60, "evidence_urls": 40,
    "contact_name": 22, "title": 26, "email": 28, "email_status": 12,
    "direct_phone": 16, "corporate_phone": 16, "linkedin": 30, "city": 16, "state": 14,
    "phone": 16, "subject": 34, "body": 68, "followup_subject": 34,
    "followup_body": 68, "talking_points": 68,
}
_XLSX_WRAP = {"fit_summary", "evidence_urls", "recommended_products", "software_resold",
              "all_verified_emails", "subject", "body", "followup_subject",
              "followup_body", "talking_points"}
_XLSX_CENTER = {"tier", "score", "total_contacts", "verified_emails", "apollo_verified",
                "validation", "employees"}


def _fill_sheet(ws, title: str, spec: list, rows: list[dict]) -> None:
    """Populate a worksheet from a (header, key) spec with the shared styling."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tier_fill = {
        "A": PatternFill("solid", fgColor="C6EFCE"), "B": PatternFill("solid", fgColor="BDD7EE"),
        "C": PatternFill("solid", fgColor="FFF2CC"), "D": PatternFill("solid", fgColor="F2F2F2"),
    }
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = title
    for col, (header, _key) in enumerate(spec, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for ri, r in enumerate(rows, 2):
        for ci, (_header, key) in enumerate(spec, 1):
            cell = ws.cell(row=ri, column=ci, value=r.get(key, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=key in _XLSX_WRAP,
                                       horizontal="center" if key in _XLSX_CENTER else "left")
            if key == "tier":
                fill = tier_fill.get(str(r.get("tier", "")).upper())
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)
    for col, (_header, key) in enumerate(spec, 1):
        ws.column_dimensions[get_column_letter(col)].width = _XLSX_WIDTHS.get(key, 16)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
